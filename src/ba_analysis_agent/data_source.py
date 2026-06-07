from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from .models import ColumnProfile, DataSourceProfile, TableProfile


SENSITIVE_TOKENS = (
    "mobile",
    "phone",
    "vin",
    "customer_id",
    "spark_id",
    "cop_id",
    "id_card",
    "identity",
    "email",
)

DOMAIN_PREFIXES = {
    "register": "register",
    "leads": "leads",
    "oppty": "oppty",
    "opportunity": "oppty",
    "visit": "visit",
    "td": "td",
    "order": "order",
    "dealer": "dealer",
    "dealership": "dealer",
    "province": "dealer",
    "city": "dealer",
    "region": "dealer",
    "sales": "dealer",
    "channel": "channel",
    "media": "channel",
    "campaign": "channel",
    "brand": "product",
    "series": "product",
    "model": "product",
    "variant": "product",
    "customer": "customer",
}


def build_profile(path: str | Path) -> DataSourceProfile:
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix == ".xlsx":
        return profile_xlsx(source_path)
    if suffix == ".csv":
        return profile_csv(source_path)
    raise ValueError(f"Unsupported source type: {suffix}. Supported: .xlsx, .csv")


def profile_xlsx(path: Path) -> DataSourceProfile:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to profile .xlsx files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        tables: list[TableProfile] = []
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            headers = [str(value).strip() for value in next(rows, ()) if value is not None]
            sample_values = [[] for _ in headers]
            for row in worksheet.iter_rows(min_row=2, max_row=min(worksheet.max_row, 101), values_only=True):
                for index, value in enumerate(row[: len(headers)]):
                    if value is not None:
                        sample_values[index].append(value)
            columns = [
                _column_profile(header, dtype=_infer_dtype(sample_values[index]))
                for index, header in enumerate(headers)
            ]
            row_count = max(worksheet.max_row - 1, 0) if worksheet.max_row else None
            tables.append(TableProfile(name=worksheet.title, row_count=row_count, columns=columns))
    finally:
        workbook.close()

    return DataSourceProfile(
        source_type="xlsx",
        source_name=str(path),
        tables_or_sheets=tables,
        business_domains=_business_domains(table.columns for table in tables),
    )


def profile_csv(path: Path) -> DataSourceProfile:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        headers = [header.strip() for header in next(reader, [])]
        sample_values = [[] for _ in headers]
        row_count = 0
        for row in reader:
            row_count += 1
            if row_count <= 100:
                for index, value in enumerate(row[: len(headers)]):
                    if value != "":
                        sample_values[index].append(value)

    table = TableProfile(
        name=path.stem,
        row_count=row_count,
        columns=[
            _column_profile(header, dtype=_infer_dtype(sample_values[index]))
            for index, header in enumerate(headers)
            if header
        ],
    )
    return DataSourceProfile(
        source_type="csv",
        source_name=str(path),
        tables_or_sheets=[table],
        business_domains=_business_domains([table.columns]),
    )


def profile_database_schema(
    source_name: str,
    tables: dict[str, Iterable[str]],
) -> DataSourceProfile:
    table_profiles = [
        TableProfile(
            name=table_name,
            row_count=None,
            columns=[_column_profile(column) for column in columns],
        )
        for table_name, columns in tables.items()
    ]
    return DataSourceProfile(
        source_type="database",
        source_name=source_name,
        tables_or_sheets=table_profiles,
        business_domains=_business_domains(table.columns for table in table_profiles),
    )


def _column_profile(name: str, dtype: str = "unknown") -> ColumnProfile:
    normalized = name.lower()
    return ColumnProfile(
        name=name,
        dtype=dtype,
        business_domain=_infer_domain(normalized),
        is_sensitive=any(token in normalized for token in SENSITIVE_TOKENS),
        sample_allowed=False,
    )


def _infer_domain(normalized_name: str) -> str:
    if any(token in normalized_name for token in ("channel", "media", "campaign", "source", "platform")):
        return "channel"
    if any(token in normalized_name for token in ("brand", "series", "model", "variant", "energy")):
        return "product"
    if any(token in normalized_name for token in ("dealer", "dealership", "city", "province", "region", "area", "investor")):
        return "dealer"
    parts = normalized_name.split("_")
    candidates = parts + [normalized_name]
    for candidate in candidates:
        if candidate in DOMAIN_PREFIXES:
            return DOMAIN_PREFIXES[candidate]
    return "general"


def _business_domains(column_groups: Iterable[Iterable[ColumnProfile]]) -> list[str]:
    domains = {column.business_domain for columns in column_groups for column in columns}
    preferred_order = [
        "register",
        "leads",
        "oppty",
        "visit",
        "td",
        "order",
        "dealer",
        "channel",
        "product",
        "customer",
        "general",
    ]
    return [domain for domain in preferred_order if domain in domains]


def _infer_dtype(values: list[object]) -> str:
    if not values:
        return "empty"

    observed: set[str] = set()
    for value in values:
        if isinstance(value, bool):
            observed.add("boolean")
        elif isinstance(value, (datetime, date)):
            observed.add("datetime")
        elif isinstance(value, int) and not isinstance(value, bool):
            observed.add("integer")
        elif isinstance(value, float):
            observed.add("number")
        else:
            text = str(value).strip()
            observed.add(_infer_text_dtype(text))

    if observed <= {"integer"}:
        return "integer"
    if observed <= {"integer", "number"}:
        return "number"
    if observed <= {"datetime"}:
        return "datetime"
    if len(observed) == 1:
        return next(iter(observed))
    return "mixed"


def _infer_text_dtype(text: str) -> str:
    if text.lower() in {"true", "false"}:
        return "boolean"
    try:
        int(text)
        return "integer"
    except ValueError:
        pass
    try:
        float(text)
        return "number"
    except ValueError:
        pass
    return "string"
