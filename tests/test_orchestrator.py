from __future__ import annotations

import unittest

from ba_analysis_agent.data_source import profile_database_schema
from ba_analysis_agent.llm import parse_json_object, safe_context_from_profile
from ba_analysis_agent.models import AgentState, Stage, StageStatus
from ba_analysis_agent.orchestrator import BAAnalysisOrchestrator, MasterAnalyticsAgent
from ba_analysis_agent.web_app import ChatAppState


class FakeLLMService:
    provider_name = "fake"

    def enhance_artifact(self, stage_id, business_problem, base_artifact, safe_context):
        if stage_id == "analysis_design":
            result = dict(base_artifact)
            result["analysis_purpose"] = f"LLM增强: {business_problem}"
            result["business_hypotheses"] = [
                {
                    "factor": "LLM生成因素",
                    "hypothesis": "基于字段元数据生成的动态假设。",
                    "why_it_matters": "验证 LLM 增强层确实生效。",
                    "validation_approach": "检查输出是否覆盖规则草稿。",
                    "evidence_status": "待验证",
                }
            ]
            return result
        return None


class FakeIntentLLMService:
    provider_name = "fake_intent"
    last_error = None

    def enhance_artifact(self, stage_id, business_problem, base_artifact, safe_context):
        if stage_id == "intent_understanding":
            return {
                "time_range": "2026-03",
                "comparison_baseline": "2026-02",
                "metric_focus": "转化率",
                "metric": "成交效率",
                "issue_type": "下降",
                "funnel_scope": {
                    "start_stage": "线索",
                    "end_stage": "订单",
                    "mentioned_stages": ["线索", "订单"],
                },
                "dimensions": ["渠道", "区域"],
                "filters": ["排除取消订单"],
                "audience": "销售管理",
                "open_questions": [],
                "confidence": 0.93,
                "understanding_summary": "LLM 已将成交变差理解为线索到订单成交效率下降。",
            }
        return None


class FakeAnalysisDesignLLMFirstService:
    provider_name = "fake_design"
    last_error = None

    def enhance_artifact(self, stage_id, business_problem, base_artifact, safe_context):
        if stage_id == "intent_understanding":
            return None
        if stage_id == "analysis_design_llm_first":
            return {
                "analysis_purpose": "LLM直接生成：诊断成交效率变化的业务原因。",
                "business_context": {
                    "purpose": "诊断成交效率变化的业务原因。",
                    "business_background": "由 LLM 基于 analysis_context 和字段元数据生成。",
                    "core_problems": ["成交效率变化是否由渠道结构或区域承接差异导致？"],
                    "related_departments": ["销售管理", "渠道投放"],
                    "key_metrics": ["成交效率", "渠道贡献度"],
                    "urgency": "中等",
                    "confidence": 0.91,
                },
                "clarification_questions": ["是否排除取消订单？"],
                "business_hypotheses": [
                    {
                        "factor": "渠道结构变化",
                        "hypothesis": "高转化渠道占比下降可能拉低成交效率。",
                        "why_it_matters": "结构变化会影响整体表现。",
                        "validation_approach": "按渠道拆解成交效率和贡献度。",
                        "related_fields": ["leads_channel_name", "leads_id", "order_id"],
                        "evidence_status": "待验证",
                    }
                ],
                "metrics_tree": [
                    {
                        "name": "成交效率",
                        "description": "LLM 生成的指标树",
                        "metrics": ["线索到订单转化率", "渠道贡献度"],
                    }
                ],
                "field_requirements": {
                    "funnel_time_fields": ["leads_create_time", "order_first_confirm_time"],
                    "funnel_id_fields": ["leads_id", "order_id"],
                    "first_flag_fields": ["leads_first_flag", "order_first_flag"],
                    "dimension_fields": ["leads_channel_name", "region_route"],
                    "metric_fields": [],
                    "missing_recommended_fields": [],
                },
                "analysis_path": ["确认时间范围和对比基准。", "按渠道和区域拆解成交效率。"],
                "confidence_score": 0.91,
            }
        return None


class OrchestratorTests(unittest.TestCase):
    def setUp(self) -> None:
        self.profile = profile_database_schema(
            "demo",
            {
                "Sheet1": [
                    "register_create_time",
                    "leads_create_time",
                    "oppty_create_time",
                    "visit_arrival_time",
                    "td_start_time",
                    "order_first_confirm_time",
                    "register_rcid",
                    "leads_id",
                    "oppty_id",
                    "visit_id",
                    "td_id",
                    "order_id",
                    "register_first_flag",
                    "leads_first_flag",
                    "oppty_first_flag",
                    "visit_first_flag",
                    "td_first_flag",
                    "order_first_flag",
                    "register_mobile_phone_number",
                    "order_vin_17",
                    "brand_route",
                    "region_route",
                    "leads_channel_name",
                ],
            },
        )

    def test_sensitive_fields_are_marked(self) -> None:
        columns = self.profile.tables_or_sheets[0].columns
        sensitive = {column.name for column in columns if column.is_sensitive}
        self.assertIn("register_mobile_phone_number", sensitive)
        self.assertIn("order_vin_17", sensitive)

    def test_database_profile_has_domains(self) -> None:
        self.assertIn("register", self.profile.business_domains)
        self.assertIn("order", self.profile.business_domains)
        self.assertIn("channel", self.profile.business_domains)

    def test_stage_gate_requires_confirmation(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        artifact = orchestrator.start("分析注册到订单转化率下降")
        self.assertEqual(artifact.stage, Stage.ANALYSIS_DESIGN)
        self.assertEqual(artifact.status, StageStatus.DRAFT)
        self.assertIsInstance(orchestrator, MasterAnalyticsAgent)
        self.assertEqual(orchestrator.status()["agent_state"], AgentState.WAITING_APPROVAL.value)
        self.assertEqual(orchestrator.status()["active_stage"], Stage.ANALYSIS_DESIGN.value)

        next_artifact = orchestrator.confirm_current()
        self.assertEqual(next_artifact.stage, Stage.DATA_PLAN)
        self.assertEqual(orchestrator.status()["stages"][Stage.ANALYSIS_DESIGN.value], StageStatus.BA_CONFIRMED.value)

    def test_analysis_design_business_hypotheses_are_factor_based(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        artifact = orchestrator.start("分析3月线索到订单转化率下降的原因")
        hypotheses = artifact.payload["business_hypotheses"]
        self.assertGreaterEqual(len(hypotheses), 3)
        self.assertIn("analysis_defaults", artifact.payload)
        self.assertNotIn("时间范围", hypotheses[0])
        self.assertIn("factor", hypotheses[0])
        self.assertIn("hypothesis", hypotheses[0])
        factors = [item["factor"] for item in hypotheses]
        self.assertIn("线索到订单链路断点", factors)
        self.assertGreater(len(artifact.payload["field_requirements"]["dimension_fields"]), 0)

    def test_analysis_design_changes_with_question_intent(self) -> None:
        funnel_agent = BAAnalysisOrchestrator(self.profile)
        funnel_artifact = funnel_agent.start("分析3月线索到订单转化率下降的原因")

        channel_agent = BAAnalysisOrchestrator(self.profile)
        channel_artifact = channel_agent.start("分析渠道线索量下降的原因")

        self.assertNotEqual(
            funnel_artifact.payload["analysis_purpose"],
            channel_artifact.payload["analysis_purpose"],
        )
        self.assertIn("线索到订单", funnel_artifact.payload["analysis_purpose"])
        self.assertIn("线索", channel_artifact.payload["analysis_purpose"])
        self.assertIn("渠道", channel_artifact.payload["analysis_context"]["dimensions"])
        self.assertNotEqual(
            funnel_artifact.payload["metrics_tree"][1]["metrics"],
            channel_artifact.payload["metrics_tree"][1]["metrics"],
        )

    def test_channel_volume_question_prioritizes_channel_volume_hypotheses(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        artifact = orchestrator.start("分析渠道线索量下降的原因")
        self.assertEqual(artifact.payload["detected_intent"]["metric_focus"], "规模")
        self.assertIn("渠道", artifact.payload["detected_intent"]["focus_dimensions"])
        self.assertIn("数量变化", artifact.payload["metrics_tree"][1]["name"])
        factors = [item["factor"] for item in artifact.payload["business_hypotheses"]]
        self.assertIn("线索供给变化", factors)
        self.assertIn("渠道流量或投放变化", factors)

    def test_playbook_guidance_matches_channel_topic(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        artifact = orchestrator.start("分析渠道线索量下降的原因")
        guidance = artifact.payload["playbook_guidance"]
        self.assertTrue(guidance["enabled"])
        topic_titles = [topic["title"] for topic in guidance["matched_topics"]]
        self.assertIn("Channel And Campaign Quality", topic_titles)
        self.assertIn("playbook_recommended_fields", artifact.payload["field_requirements"])
        self.assertTrue(any(item["factor"].startswith("Playbook主题") for item in artifact.payload["business_hypotheses"]))

    def test_playbook_guidance_matches_dealer_region_topic(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        artifact = orchestrator.start("分析区域订单转化异常的原因")
        topic_titles = [topic["title"] for topic in artifact.payload["playbook_guidance"]["matched_topics"]]
        self.assertIn("Dealer And Region Execution", topic_titles)

    def test_llm_enhancement_can_override_design_content(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile, llm_service=FakeLLMService())
        artifact = orchestrator.start("分析3月线索到订单转化率下降的原因")
        self.assertTrue(artifact.payload["analysis_purpose"].startswith("LLM增强: 分析3月线索到订单转化率下降的原因"))
        self.assertEqual(artifact.payload["llm_enrichment"]["status"], "applied")
        self.assertEqual(artifact.payload["business_hypotheses"][0]["factor"], "LLM生成因素")

    def test_llm_safe_context_contains_metadata_only(self) -> None:
        context = safe_context_from_profile(self.profile)
        first_column = context["tables_or_sheets"][0]["columns"][0]
        self.assertIn("name", first_column)
        self.assertIn("dtype", first_column)
        self.assertIn("is_sensitive", first_column)
        self.assertNotIn("examples", first_column)
        self.assertNotIn("sample_rows", context["tables_or_sheets"][0])

    def test_llm_json_parser_accepts_markdown_json_fence(self) -> None:
        parsed = parse_json_object('```json\n{"analysis_purpose": "ok"}\n```')
        self.assertEqual(parsed, {"analysis_purpose": "ok"})

    def test_llm_json_parser_skips_unclosed_think_prefix(self) -> None:
        parsed = parse_json_object('<think>reasoning text {"ignored": true}\n{"analysis_purpose": "ok"}')
        self.assertEqual(parsed, {"analysis_purpose": "ok"})

    def test_data_plan_uses_existing_fields(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("分析注册到订单转化率下降")
        data_plan = orchestrator.confirm_current()
        sql = data_plan.payload["sql_plan"]["sql_queries"][0]["sql"]
        self.assertIn("register_rcid", sql)
        self.assertIn("order_id", sql)
        self.assertNotIn("fake_field", sql)

    def test_revision_does_not_advance_stage(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("分析注册到订单转化率下降")
        revised = orchestrator.request_revision("增加渠道拆解")
        self.assertEqual(revised.status, StageStatus.DRAFT)
        self.assertEqual(orchestrator.status()["active_stage"], Stage.ANALYSIS_DESIGN.value)
        self.assertEqual(orchestrator.status()["revision_count"], 1)
        self.assertEqual(revised.payload["ba_feedback"][0]["type"], "revision")
        self.assertIn("渠道", revised.payload["detected_intent"]["focus_dimensions"])

    def test_ba_feedback_is_recorded_without_advancing_stage(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("分析3月线索到订单转化率下降的原因")
        updated = orchestrator.add_ba_feedback("时间范围是3月，对比2月，按渠道和区域拆解")
        self.assertEqual(updated.stage, Stage.ANALYSIS_DESIGN)
        self.assertEqual(updated.status, StageStatus.DRAFT)
        self.assertEqual(orchestrator.status()["active_stage"], Stage.ANALYSIS_DESIGN.value)
        self.assertEqual(updated.payload["ba_feedback"][0]["type"], "clarification")
        self.assertIn("按渠道和区域拆解", updated.payload["ba_feedback"][0]["text"])
        self.assertEqual(updated.payload["detected_intent"]["time_period"], "3月")
        self.assertIn("渠道", updated.payload["detected_intent"]["focus_dimensions"])
        self.assertIn("区域", updated.payload["detected_intent"]["focus_dimensions"])

    def test_negative_dimension_feedback_excludes_dimension(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("帮我看看3月为什么成交变差了")
        updated = orchestrator.add_ba_feedback("成交指线索到订单转化率，对比2月，先按渠道和区域看，不看车型")
        dimensions = updated.payload["analysis_context"]["dimensions"]
        self.assertIn("渠道", dimensions)
        self.assertIn("区域", dimensions)
        self.assertNotIn("产品", dimensions)

    def test_intent_understanding_uses_llm_when_available(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile, llm_service=FakeIntentLLMService())
        artifact = orchestrator.start("帮我看看3月为什么成交变差了")
        context = artifact.payload["analysis_context"]
        self.assertEqual(context["time_range"], "2026-03")
        self.assertEqual(context["comparison_baseline"], "2026-02")
        self.assertEqual(context["metric"], "成交效率")
        self.assertEqual(context["dimensions"], ["渠道", "区域"])
        self.assertEqual(context["llm_enrichment"]["status"], "applied")

    def test_analysis_design_uses_llm_first_without_playbook_when_available(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile, llm_service=FakeAnalysisDesignLLMFirstService())
        artifact = orchestrator.start("帮我看看3月为什么成交变差了")
        self.assertEqual(artifact.payload["generation_mode"], "llm_first_no_skill")
        self.assertEqual(artifact.payload["skill"], "llm_generated")
        self.assertEqual(artifact.payload["analysis_purpose"], "LLM直接生成：诊断成交效率变化的业务原因。")
        self.assertNotIn("playbook_guidance", artifact.payload)
        self.assertEqual(artifact.payload["llm_enrichment"]["status"], "applied")

    def test_history_records_operations(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("分析注册到订单转化率下降")
        orchestrator.confirm_current()
        history = orchestrator.history()
        operation_types = [item.operation_type for item in history["operations"]]
        self.assertIn("workflow_started", operation_types)
        self.assertIn("stage_approved", operation_types)
        self.assertIn("approval_requested", operation_types)

    def test_workflow_config_is_exposed(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        workflow = orchestrator.workflow()
        self.assertEqual(workflow.workspace_type, "analytics")
        self.assertEqual(workflow.first_stage(), Stage.ANALYSIS_DESIGN)

    def test_report_plan_waits_for_insight_confirmation(self) -> None:
        orchestrator = BAAnalysisOrchestrator(self.profile)
        orchestrator.start("分析注册到订单转化率下降")
        orchestrator.confirm_current()
        insight_review = orchestrator.confirm_current()
        self.assertEqual(insight_review.stage, Stage.INSIGHT_REVIEW)

        report_plan = orchestrator.confirm_current()
        self.assertEqual(report_plan.stage, Stage.REPORT_PLAN)


class WebChatTests(unittest.TestCase):
    def test_web_chat_message_starts_analysis(self) -> None:
        import tempfile
        from pathlib import Path
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "demo.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["leads_create_time", "order_first_confirm_time", "leads_id", "order_id", "leads_channel_name"])
            sheet.append(["2026-03-01", "2026-03-05", "L1", "O1", "官方渠道"])
            workbook.save(path)

            app = ChatAppState(path)
            response = app.handle_message("分析渠道线索量下降的原因")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["stage"], Stage.ANALYSIS_DESIGN.value)
            self.assertIn("detected_intent", response["data"]["payload"])
            self.assertIn("view", response)
            self.assertIn("summary", response["view"])
            self.assertGreater(len(response["view"]["sections"]), 0)
            self.assertTrue(any(section["title"] == "业务假设" for section in response["view"]["sections"]))

            feedback_response = app.handle_message("时间范围是3月，对比2月，按渠道和区域拆解")
            self.assertTrue(feedback_response["ok"])
            self.assertEqual(feedback_response["data"]["stage"], Stage.ANALYSIS_DESIGN.value)
            self.assertEqual(feedback_response["data"]["payload"]["ba_feedback"][0]["type"], "clarification")
            self.assertTrue(any(section["title"] == "BA 已补充/修改" for section in feedback_response["view"]["sections"]))

    def test_web_chat_can_enable_openai_for_current_process(self) -> None:
        import tempfile
        from pathlib import Path
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "demo.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["leads_create_time", "order_first_confirm_time", "leads_id", "order_id"])
            workbook.save(path)

            app = ChatAppState(path)
            response = app.configure_openai("sk-test", "gpt-4.1-mini")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["provider"], "openai_responses")
            self.assertTrue(response["data"]["enabled"])

    def test_web_chat_can_enable_minimax_for_current_process(self) -> None:
        import tempfile
        from pathlib import Path
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "demo.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["leads_create_time", "order_first_confirm_time", "leads_id", "order_id"])
            workbook.save(path)

            app = ChatAppState(path)
            response = app.configure_llm("minimax", "minimax-test-token", "MiniMax-M2.7")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["provider"], "minimax")
            self.assertTrue(response["data"]["enabled"])

            response = app.configure_llm("minimax_cn", "minimax-test-token", "MiniMax-M2.7")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["provider"], "minimax_cn")
            self.assertTrue(response["data"]["enabled"])

            response = app.configure_llm("minimax_anthropic", "minimax-test-token", "MiniMax-M2.7")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["provider"], "minimax_anthropic")
            self.assertTrue(response["data"]["enabled"])

            response = app.configure_llm("minimax_anthropic_sdk", "minimax-test-token", "MiniMax-M2.7")
            self.assertTrue(response["ok"])
            self.assertEqual(response["data"]["provider"], "minimax_anthropic_sdk")
            self.assertTrue(response["data"]["enabled"])

    def test_web_chat_rejects_non_ascii_llm_token(self) -> None:
        import tempfile
        from pathlib import Path
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "demo.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["leads_create_time", "order_first_confirm_time", "leads_id", "order_id"])
            workbook.save(path)

            app = ChatAppState(path)
            response = app.configure_llm("minimax", "这是token", "MiniMax-M2.7")
            self.assertFalse(response["ok"])
            self.assertIn("非 ASCII", response["text"])


if __name__ == "__main__":
    unittest.main()
